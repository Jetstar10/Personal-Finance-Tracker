from email.header import decode_header
from email.utils import parsedate_to_datetime
from pathlib import Path

from datetime import datetime
import email
import html
import imaplib
import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parent
IMAP_SERVER = "imap.gmail.com"
IMAP_PORT = 993
BANK_SENDER = "ibanking.alert@dbs.com"
OUTPUT_DIR = BASE_DIR / "financial_reports"
MONTHLY_WORKBOOK = OUTPUT_DIR / "monthly_financials.xlsx"
LAST_UID_FILE = BASE_DIR / "last_uid.txt"
ACCOUNT_ENDING = os.environ.get("JARVIS_DBS_ACCOUNT_ENDING", "5209")
MONTH_ORDER = {
    "January": 1,
    "February": 2,
    "March": 3,
    "April": 4,
    "May": 5,
    "June": 6,
    "July": 7,
    "August": 8,
    "September": 9,
    "October": 10,
    "November": 11,
    "December": 12,
    "unknown": 99,
}
MONTH_NAMES = {
    "jan": "January",
    "feb": "February",
    "mar": "March",
    "apr": "April",
    "may": "May",
    "jun": "June",
    "jul": "July",
    "aug": "August",
    "sep": "September",
    "oct": "October",
    "nov": "November",
    "dec": "December",
}

TRANSACTION_HEADERS = [
    "UID",
    "Transaction Date",
    "Email Date",
    "Type",
    "Category",
    "Party",
    "Description",
    "Amount",
    "Income",
    "Expense",
    "Net",
    "Source Account",
    "Destination",
    "Subject",
]
UNPARSED_HEADERS = ["UID", "Transaction Date", "Party", "Detected Amount", "Category Guess", "Subject"]

CATEGORY_RULES = {
    "Investment": ["singapore government securities", "t-bill", "treasury bill", "sgs", "bond"],
    "Salary": ["salary", "payroll", "wages"],
    "Refund": ["refund", "reversal", "cashback", "rebate"],
    "Food Delivery": ["grabfood", "foodpanda", "deliveroo"],
    "Dining": ["mcdonald", "starbucks", "kopitiam", "restaurant", "cafe", "coffee", "toast box", "ya kun", "kfc", "burger king", "subway"],
    "Groceries": ["fairprice", "ntuc", "sheng siong", "cold storage", "giant", "supermarket", "grocery"],
    "Ride Hailing": ["grab", "gojek", "tada", "comfortdelgro", "taxi"],
    "Public Transport": ["mrt", "ez-link", "simplygo", "transitlink", "bus/mrt"],
    "Online Shopping": ["shopee", "lazada", "amazon", "qoo10"],
    "Retail Shopping": ["retail", "purchase", "uniqlo", "courts", "challenger", "watsons", "guardian"],
    "Utilities": ["utilities", "sp services", "electricity", "water"],
    "Telecom": ["singtel", "starhub", "m1", "circles.life", "gomo", "simba"],
    "Insurance": ["insurance", "aia", "prudential", "great eastern", "income insurance"],
    "Medical": ["clinic", "hospital", "pharmacy", "doctor", "dental"],
    "Subscriptions": ["subscription", "netflix", "spotify", "apple.com/bill", "google", "youtube"],
    "Transfer": ["paynow", "fund transfer", "transfer", "fast payment"],
    "Bills": ["bill"],
    "Banking": ["atm", "cash withdrawal", "bank charge", "service fee"],
}


def decode_mime_words(value):
    if not value:
        return ""

    decoded = []
    for part, encoding in decode_header(value):
        if isinstance(part, bytes):
            decoded.append(part.decode(encoding or "utf-8", errors="replace"))
        else:
            decoded.append(part)
    return "".join(decoded)


def html_to_text(html_text):
    html_text = re.sub(r"(?is)<script.*?>.*?</script>", " ", html_text)
    html_text = re.sub(r"(?is)<style.*?>.*?</style>", " ", html_text)
    html_text = re.sub(r"(?s)<[^>]+>", " ", html_text)
    html_text = html.unescape(html_text)
    return re.sub(r"\s+", " ", html_text).strip()


def extract_email_body(msg):
    plain_text = ""
    html_text = ""
    parts = msg.walk() if msg.is_multipart() else [msg]

    for part in parts:
        if "attachment" in str(part.get("Content-Disposition", "")).lower():
            continue

        payload = part.get_payload(decode=True)
        if not payload:
            continue

        charset = part.get_content_charset() or "utf-8"
        decoded = payload.decode(charset, errors="replace")
        if part.get_content_type() == "text/plain" and not plain_text:
            plain_text = decoded
        elif part.get_content_type() == "text/html" and not html_text:
            html_text = decoded

    if plain_text.strip():
        return plain_text.strip()
    if html_text.strip():
        return html_to_text(html_text)
    return ""


def clean_field(value):
    return re.sub(r"\s+", " ", value or "").strip()


def email_date_details(msg):
    raw_date = decode_mime_words(msg.get("Date", ""))
    try:
        parsed = parsedate_to_datetime(raw_date)
    except (TypeError, ValueError, IndexError):
        parsed = None

    if parsed is None:
        return {"display": raw_date, "timestamp": 0}

    return {
        "display": parsed.strftime("%Y-%m-%d %H:%M:%S %z").strip(),
        "timestamp": parsed.timestamp(),
    }


def transaction_period(transaction_date):
    transaction_date = clean_field(transaction_date)
    month = "unknown"
    year = "unknown"

    month_match = re.search(r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\b", transaction_date, re.IGNORECASE)
    if month_match:
        month = MONTH_NAMES[month_match.group(1).lower()[:3]]

    year_match = re.search(r"\b(19\d{2}|20\d{2})\b", transaction_date)
    if year_match:
        year = year_match.group(1)

    return {"year": year, "month": month, "key": f"{year}_{month}"}


def transaction_month_number(transaction_date):
    month = transaction_period(transaction_date)["month"]
    if month == "unknown":
        return None
    return MONTH_ORDER.get(month)


def transaction_explicit_year(transaction_date):
    match = re.search(r"\b(19\d{2}|20\d{2})\b", clean_field(transaction_date))
    return int(match.group(1)) if match else None


def parse_email_datetime(value):
    email_date = clean_field(str(value))
    if not email_date:
        return None

    try:
        return parsedate_to_datetime(email_date)
    except (TypeError, ValueError, IndexError):
        pass

    for date_format in ("%Y-%m-%d %H:%M:%S %z", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(email_date, date_format)
        except ValueError:
            pass

    return None


def email_period(email_date):
    parsed = parse_email_datetime(email_date)
    if not parsed:
        return None

    month = MONTH_NAMES[parsed.strftime("%b").lower()]
    return {"year": str(parsed.year), "month": month, "key": f"{parsed.year}_{month}"}


def report_month_for_item(item):
    transaction_month = transaction_period(item.get("transaction_date"))["month"]
    if transaction_month != "unknown":
        return transaction_month

    fallback = email_period(item.get("email_date"))
    if fallback:
        return fallback["month"]

    return "unknown"


def period_for_year_month(year, month):
    return {"year": str(year), "month": month or "unknown", "key": f"{year}_{month or 'unknown'}"}


def short_month(month):
    return month[:3] if month != "unknown" else "unknown"


def sheet_title(period, suffix):
    base = f"{period['year']}_{short_month(period['month'])}_{suffix}"
    title = re.sub(r"[:\\/?*\[\]]+", "_", base)
    return title[:31]


def period_sort_key(period):
    year = 9999 if period["year"] == "unknown" else int(period["year"])
    return (year, MONTH_ORDER.get(period["month"], 99))


def is_my_account(text):
    normalized = clean_field(text).lower().replace(" ", "")
    markers = [
        f"yourdbs/posbaccountending{ACCOUNT_ENDING}",
        f"dbs/posbaccountending{ACCOUNT_ENDING}",
        f"myaccounta/cending{ACCOUNT_ENDING}",
        f"myaccountending{ACCOUNT_ENDING}",
        f"accountending{ACCOUNT_ENDING}",
    ]
    return any(marker in normalized for marker in markers)


def clean_party_name(text):
    text = clean_field(text)
    if not text:
        return ""

    text = re.sub(r"\s*\([^)]*\)", "", text).strip()
    if is_my_account(text):
        return "DBS/POSB"

    text = re.sub(r"\bA/C\b.*$", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\baccount ending\s+\d+.*$", "", text, flags=re.IGNORECASE).strip()
    return text


def first_match(patterns, text):
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return clean_field(match.group(1))
    return ""


def extract_amount(flat_text):
    amount_patterns = [
        r"Amount:\s*SGD\s*([\d,]+(?:\.\d{2})?)",
        r"received\s+SGD\s*([\d,]+(?:\.\d{2})?)",
        r"\bSGD\s*([\d,]+(?:\.\d{2})?)\b",
        r"\bS\$\s*([\d,]+(?:\.\d{2})?)\b",
        r"\$\s*([\d,]+(?:\.\d{2})?)\b",
    ]
    for pattern in amount_patterns:
        match = re.search(pattern, flat_text, re.IGNORECASE)
        if match:
            return float(match.group(1).replace(",", ""))
    return None


def infer_category(text, transaction_type):
    haystack = clean_field(text).lower()
    for category, keywords in CATEGORY_RULES.items():
        if any(keyword in haystack for keyword in keywords):
            return category
    if transaction_type == "Income":
        return "Income"
    if transaction_type == "Expense":
        return "Uncategorised Expense"
    return "Uncategorised"


def reporting_category(category):
    return category or "Uncategorised"


def is_investment_category(category):
    return category == "Investment"


def percent_change(current, previous):
    if previous in ("", None) or float(previous or 0) == 0:
        return ""
    return (float(current or 0) - float(previous)) / abs(float(previous))


def excel_percent_change_formula(month_cell, current_cell, previous_cell):
    return f'=IF({month_cell}="January","Nil",IF({previous_cell}=0,"Nil",({current_cell}-{previous_cell})/{previous_cell}))'


def money_formula(value):
    return f"={float(value or 0):.2f}"


def build_description(subject, parsed, body_text):
    pieces = [
        parsed.get("type", ""),
        parsed.get("party", ""),
        f"SGD {parsed.get('amount'):.2f}" if isinstance(parsed.get("amount"), (int, float)) else "",
        parsed.get("transaction_date", ""),
        subject,
    ]
    description = " | ".join(piece for piece in pieces if piece)
    return description or clean_field(body_text)[:500]


def force_investment_as_expense(item):
    if item.get("category") != "Investment":
        return item

    amount = item.get("amount") or item.get("expense") or item.get("income")
    if amount not in ("", None):
        amount = float(str(amount).replace(",", ""))

    item["type"] = "Expense"
    item["income"] = ""
    item["expense"] = amount if amount not in ("", None) else ""
    item["net"] = -amount if amount not in ("", None) else ""
    if not item.get("party"):
        item["party"] = item.get("destination") or item.get("source_account") or "Investment"
    return item


def refine_item_category(item):
    current = item.get("category") or "Uncategorised"
    if current == "Investment":
        return force_investment_as_expense(item)

    text = " ".join(
        clean_field(str(item.get(field) or ""))
        for field in ("subject", "description", "party", "source_account", "destination")
    )
    inferred = infer_category(text, item.get("type"))
    generic_categories = {"Uncategorised", "Uncategorised Expense", "Transfer", "Banking", "Bills", "Shopping", "Transport", "Food"}
    if inferred not in generic_categories or current in generic_categories:
        item["category"] = inferred

    return force_investment_as_expense(item)


def parse_transaction(text, subject=""):
    result = {
        "transaction_date": "",
        "type": "Unknown",
        "amount": "",
        "income": "",
        "expense": "",
        "source_account": "",
        "destination": "",
        "party": "",
        "category": "Uncategorised",
        "description": "",
    }
    if not text:
        return result

    flat_text = re.sub(r"\s+", " ", text.replace("\r", "\n")).strip()
    lower_text = flat_text.lower()
    amount = extract_amount(flat_text)
    result["amount"] = amount if amount is not None else ""

    transaction_date = first_match(
        [
            r"Date\s*&\s*Time:\s*([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{4}\s+[0-9]{2}:[0-9]{2}\s*\([A-Za-z]+\))",
            r"Date\s*&\s*Time:\s*([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{4}\s+[0-9]{2}:[0-9]{2})",
            r"Date\s*&\s*Time:\s*([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{2}:[0-9]{2}\s*\([A-Za-z]+\))",
            r"Date\s*&\s*Time:\s*([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{2}:[0-9]{2})",
            r"on\s+([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{4}\s+[0-9]{2}:[0-9]{2}\s*[A-Za-z]*)",
        ],
        flat_text,
    )
    result["transaction_date"] = transaction_date

    raw_from = first_match(
        [r"From:\s*(.+?)(?=\s+To:|\s+Date(?:\s*(?:&|and)\s*Time)?:|\s+Amount:|$)"],
        flat_text,
    )
    raw_to = first_match(
        [
            r"To:\s*(.+?)(?=\s+From:|\s+Date(?:\s*(?:&|and)\s*Time)?:|\s+Amount:|\s+Didn[â€™']t expect|\s+If unauthorised|\s+Thank you|\s+Yours faithfully|$)"
        ],
        flat_text,
    )

    if is_my_account(raw_from):
        result["type"] = "Expense"
    elif is_my_account(raw_to):
        result["type"] = "Income"
    else:
        outgoing_keywords = ["paid", "spent", "purchase", "debited", "transfer to", "bill payment", "paynow dated"]
        incoming_keywords = ["you have received", "received", "credited", "credit to", "refund", "reversal", "transfer from", "salary", "incoming"]
        expense_phrases = [
            "debited from your",
            "deducted from your",
            "sent sgd",
            "paid sgd",
            "payment to",
            "transfer to",
            "purchase of",
        ]
        income_phrases = [
            "credited to your",
            "deposited into your",
            "received sgd",
            "you have received",
            "transfer from",
        ]
        if any(phrase in lower_text for phrase in income_phrases) or any(word in lower_text for word in incoming_keywords):
            result["type"] = "Income"
        elif any(phrase in lower_text for phrase in expense_phrases) or any(word in lower_text for word in outgoing_keywords):
            result["type"] = "Expense"

    result["source_account"] = clean_party_name(raw_from)
    result["destination"] = clean_party_name(raw_to)

    if result["type"] == "Income" and not result["source_account"]:
        result["source_account"] = clean_party_name(
            first_match(
                [
                    r"received\s+SGD\s*[\d,]+(?:\.\d{2})?\s+from\s+(.+?)(?=\s+on\s|\s+via\s|\s+ref(?:erence)?[:\s]|$)",
                    r"transfer from\s+(.+?)(?=\s+on\s|\s+via\s|\s+ref(?:erence)?[:\s]|$)",
                    r"Sender(?:'s)?\s*Name:\s*(.+?)(?=\s+[A-Za-z][A-Za-z\s/&-]*:|$)",
                    r"From\s+(.+?)(?=\s+to\s+DBS/POSB|\s+to\s+your|\s+on\s|\s+Date|\s+Amount|$)",
                ],
                flat_text,
            )
        )

    if result["type"] == "Expense":
        result["party"] = result["destination"]
        result["expense"] = amount if amount is not None else ""
    elif result["type"] == "Income":
        result["party"] = result["source_account"]
        result["income"] = amount if amount is not None else ""
    else:
        result["party"] = result["source_account"] or result["destination"]

    if amount is None or result["type"] == "Unknown":
        result["type"] = "Unknown"

    result["category"] = infer_category(f"{subject} {flat_text} {result['party']}", result["type"])
    force_investment_as_expense(result)
    result["description"] = build_description(subject, result, flat_text)
    return result


def load_last_uid():
    if not LAST_UID_FILE.exists():
        return 0
    content = LAST_UID_FILE.read_text(encoding="utf-8").strip()
    return int(content) if content.isdigit() else 0


def save_last_uid(uid):
    LAST_UID_FILE.write_text(str(uid), encoding="utf-8")


def connect_to_gmail():
    email_address = os.environ.get("JARVIS_EMAIL")
    email_password = os.environ.get("JARVIS_EMAIL_PASSWORD")
    if not email_address or not email_password:
        raise RuntimeError("Missing email credentials. Set JARVIS_EMAIL and JARVIS_EMAIL_PASSWORD first.")

    mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
    mail.login(email_address, email_password)
    mail.select("INBOX")
    return mail


def search_bank_uids(mail):
    query = f"from:{BANK_SENDER} newer_than:730d"
    status, data = mail.uid("SEARCH", None, "X-GM-RAW", f'"{query}"')
    if status == "OK" and data and data[0]:
        return data[0].split()

    status, data = mail.uid("SEARCH", None, "FROM", BANK_SENDER)
    if status == "OK" and data and data[0]:
        return data[0].split()
    return []


def fetch_new_bank_emails():
    mail = connect_to_gmail()
    last_uid = load_last_uid()
    records = []
    unparsed_records = []
    max_uid_seen = last_uid

    try:
        uid_ints = sorted(int(uid.decode() if isinstance(uid, bytes) else uid) for uid in search_bank_uids(mail))
        new_uids = [uid for uid in uid_ints if uid > last_uid]

        for uid in new_uids:
            max_uid_seen = max(max_uid_seen, uid)
            status, msg_data = mail.uid("FETCH", str(uid), "(RFC822)")
            if status != "OK" or not msg_data or msg_data[0] is None:
                continue

            msg = email.message_from_bytes(msg_data[0][1])
            subject = decode_mime_words(msg.get("Subject", ""))
            email_date = email_date_details(msg)
            body = extract_email_body(msg)
            parsed = parse_transaction(f"{subject}\n{body}", subject=subject)

            if parsed["type"] == "Unknown":
                if parsed["amount"] not in ("", None):
                    unparsed_records.append(
                        {
                            "uid": uid,
                            "transaction_date": parsed["transaction_date"],
                            "party": parsed["party"],
                            "amount": parsed["amount"],
                            "category": parsed["category"],
                            "subject": subject,
                            "email_date": email_date["display"],
                            "email_timestamp": email_date["timestamp"],
                        }
                    )
                continue

            records.append(
                {
                    "uid": uid,
                    "subject": subject,
                    "email_date": email_date["display"],
                    "email_timestamp": email_date["timestamp"],
                    "parsed": parsed,
                }
            )
    finally:
        mail.logout()

    return records, unparsed_records, max_uid_seen


def ensure_sheet(wb, name, headers):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)

    existing_headers = [ws.cell(row=1, column=col).value for col in range(1, len(headers) + 1)]
    if ws.max_row <= 1 and not any(existing_headers):
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
    return ws


def migrate_sheet_headers(ws, target_headers, row_mapper):
    current_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    if current_headers[: len(target_headers)] == target_headers:
        return

    old_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = {
            current_headers[col - 1]: ws.cell(row=row, column=col).value
            for col in range(1, len(current_headers) + 1)
            if current_headers[col - 1]
        }
        if any(value not in ("", None) for value in row_data.values()):
            old_rows.append(row_mapper(row_data))

    ws.delete_rows(1, ws.max_row or 1)
    ws.append(target_headers)
    for row in old_rows:
        ws.append(row)


def map_transaction_row(row_data):
    amount = row_data.get("Amount", "")
    income = row_data.get("Income", "")
    expense = row_data.get("Expense", "")
    if amount in ("", None):
        amount = income if income not in ("", None) else expense
    net = row_data.get("Net", "")
    if net in ("", None):
        net = float(income or 0) - float(expense or 0)
    transaction_type = "Income" if income not in ("", None) else "Expense" if expense not in ("", None) else row_data.get("Type", "")
    party = row_data.get("Party", "")
    subject = row_data.get("Subject", "")
    category = row_data.get("Category") or infer_category(f"{party} {subject}", transaction_type)
    description = row_data.get("Description") or " | ".join(str(value) for value in [transaction_type, party, subject] if value)
    return [
        row_data.get("UID", ""),
        row_data.get("Transaction Date", ""),
        row_data.get("Email Date", ""),
        transaction_type,
        category,
        party,
        description,
        amount,
        income,
        expense,
        net,
        row_data.get("Source Account", ""),
        row_data.get("Destination", ""),
        subject,
    ]


def map_unparsed_row(row_data):
    return [
        row_data.get("UID", ""),
        row_data.get("Transaction Date", ""),
        row_data.get("Party", ""),
        row_data.get("Detected Amount", ""),
        row_data.get("Category Guess", "") or infer_category(row_data.get("Subject", ""), "Unknown"),
        row_data.get("Subject", ""),
    ]


def load_report_workbook(path):
    OUTPUT_DIR.mkdir(exist_ok=True)
    if path.exists():
        return load_workbook(path)

    wb = Workbook()
    wb.active.title = "Index"
    return wb


def reset_sheet(ws, headers=None):
    ws.delete_rows(1, ws.max_row or 1)
    if headers:
        ws.append(headers)


def ensure_report_sheet(wb, name, headers=None):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)

    if headers:
        current_headers = [ws.cell(row=1, column=col).value for col in range(1, max(ws.max_column, len(headers)) + 1)]
        if ws.max_row <= 1 and not any(current_headers):
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
        elif current_headers[: len(headers)] != headers:
            mapper = map_transaction_row if headers == TRANSACTION_HEADERS else map_unparsed_row
            migrate_sheet_headers(ws, headers, mapper)
    return ws


def clear_summary(ws):
    ws.delete_rows(1, ws.max_row or 1)


def write_summary(ws, transactions_ws, period):
    clear_summary(ws)
    last_row = transactions_ws.max_row
    if transactions_ws.cell(row=last_row, column=1).value == "TOTAL":
        last_row -= 1
    last_row = max(last_row, 2)
    txn_title = transactions_ws.title.replace("'", "''")
    for row in [
        ("Monthly Performance", ""),
        ("Month", period["month"]),
        ("Year", period["year"]),
        ("", ""),
        ("Metric", "Amount"),
        ("Total Income", f"=SUM('{txn_title}'!I2:I{last_row})"),
        ("Total Expenses", f"=SUMIFS('{txn_title}'!J2:J{last_row},'{txn_title}'!E2:E{last_row},\"<>Investment\")"),
        ("Total Investments", f"=SUMIFS('{txn_title}'!J2:J{last_row},'{txn_title}'!E2:E{last_row},\"Investment\")"),
        ("Net Income", "=B6-B7"),
        ("Net After Investments", "=B6-B7-B8"),
        ("", ""),
        ("Income By Category", ""),
        ("Category", "Amount"),
    ]:
        ws.append(row)

    categories = sorted(
        {
            transactions_ws.cell(row=row, column=5).value
            for row in range(2, transactions_ws.max_row + 1)
            if transactions_ws.cell(row=row, column=5).value
        }
    )
    for category in categories:
        ws.append((category, f"=SUMIF('{txn_title}'!E:E,A{ws.max_row + 1},'{txn_title}'!I:I)"))

    ws.append(("", ""))
    ws.append(("Expense By Category", ""))
    ws.append(("Category", "Amount"))
    for category in categories:
        ws.append((category, f"=SUMIF('{txn_title}'!E:E,A{ws.max_row + 1},'{txn_title}'!J:J)"))

    ws.append(("", ""))
    ws.append(("Expenses In Parsed Order", "", "", "", ""))
    ws.append(("Transaction Date", "Party", "Category", "Amount", "Description"))
    for row in range(2, transactions_ws.max_row + 1):
        if transactions_ws.cell(row=row, column=1).value == "TOTAL":
            continue
        expense = transactions_ws.cell(row=row, column=10).value
        if expense in ("", None, 0):
            continue
        ws.append(
            (
                transactions_ws.cell(row=row, column=2).value,
                transactions_ws.cell(row=row, column=6).value,
                transactions_ws.cell(row=row, column=5).value,
                expense,
                transactions_ws.cell(row=row, column=7).value,
            )
        )


def remove_tables(ws):
    for table_name in list(ws.tables):
        del ws.tables[table_name]


def email_sort_value(value):
    parsed = parse_email_datetime(value)
    return parsed.timestamp() if parsed else 0


def remove_transaction_totals(ws):
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=1).value == "TOTAL" or not any(ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)):
            ws.delete_rows(row)


def sort_transactions_by_email_date(ws):
    remove_transaction_totals(ws)
    rows = [
        [ws.cell(row=row, column=col).value for col in range(1, len(TRANSACTION_HEADERS) + 1)]
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=1).value not in ("", None, "TOTAL")
    ]
    rows.sort(key=lambda row: (email_sort_value(row[2]), int(row[0]) if str(row[0]).isdigit() else 0), reverse=True)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append(row)


def add_transaction_totals(ws):
    sort_transactions_by_email_date(ws)
    first_data_row = 2
    last_data_row = ws.max_row
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=8, value="Net Sum / Loss")
    ws.cell(row=total_row, column=9, value=f"=SUM(I{first_data_row}:I{last_data_row})")
    ws.cell(row=total_row, column=10, value=f"=SUM(J{first_data_row}:J{last_data_row})")
    ws.cell(row=total_row, column=11, value=f"=SUM(K{first_data_row}:K{last_data_row})")


def apply_template(wb):
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    body_fill = PatternFill("solid", fgColor="F7FBFC")
    money_fill = PatternFill("solid", fgColor="E2F0D9")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                is_percent_column = (
                    (ws.title == "Yearly_Summary" and cell.column in (3, 5, 8))
                    or (ws.title == "Monthly_Summaries" and cell.column in (4, 6, 10, 11, 12))
                    or (ws.title == "Categories" and cell.column in (7, 8))
                    or (ws.title == "Chart_helper" and cell.coordinate in ("E2", "B4", "C4"))
                )
                is_money_column = (
                    ("Txn" in ws.title and cell.column in (8, 9, 10, 11))
                    or (ws.title.endswith("_Summary") and cell.column in (2, 4))
                    or (ws.title == "Yearly_Summary" and cell.column in (2, 4, 6, 7, 9))
                    or (ws.title == "Monthly_Summaries" and cell.column in (2, 3, 5, 7, 8, 9, 13))
                    or (ws.title == "Categories" and cell.column in (3, 4, 5, 6))
                    or (ws.title == "Chart_helper" and cell.column in (2, 3, 4, 6, 7, 8))
                )
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                elif is_percent_column:
                    cell.fill = money_fill
                    cell.number_format = "0.00%"
                elif is_money_column:
                    cell.fill = money_fill
                    cell.number_format = '"SGD" #,##0.00'
                else:
                    cell.fill = body_fill
        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = ws.dimensions

    for ws in wb.worksheets:
        if "Txn" in ws.title or ws.title.startswith("AllTxn_"):
            sheet_widths = {
                "A": 13,
                "B": 26,
                "C": 28,
                "D": 12,
                "E": 24,
                "F": 34,
                "G": 74,
                "H": 16,
                "I": 16,
                "J": 16,
                "K": 16,
                "L": 34,
                "M": 34,
                "N": 74,
            }
        elif "Unparsed" in ws.title:
            sheet_widths = {"A": 13, "B": 26, "C": 34, "D": 18, "E": 24, "F": 74}
        else:
            sheet_widths = {
                "A": 34,
                "B": 18,
                "C": 20,
                "D": 18,
                "E": 20,
                "F": 18,
                "G": 18,
                "H": 22,
                "I": 24,
                "J": 18,
            }

        for column, width in sheet_widths.items():
            ws.column_dimensions[column].width = width

    for ws in wb.worksheets:
        remove_tables(ws)


def periods_from_monthly_workbook(wb):
    periods = []
    for sheet_name in wb.sheetnames:
        match = re.match(r"(.+)_([A-Za-z]+)_Txn$", sheet_name)
        if not match:
            continue
        year, month_short = match.groups()
        month = next((name for name in MONTH_ORDER if short_month(name) == month_short), "unknown")
        periods.append({"year": year, "month": month, "key": f"{year}_{month}"})
    return periods


def write_index(wb):
    ws = ensure_report_sheet(wb, "Index")
    reset_sheet(ws)
    ws.append(["Financial Report Workbook", ""])
    ws.append(["Generated Sheets", ""])
    ws.append(["", ""])
    ws.append(["Monthly Sheets", "", "", "", ""])
    ws.append(["Year", "Month", "Summary Sheet", "Transaction Sheet", "Unparsed Sheet"])

    for period in sorted(periods_from_monthly_workbook(wb), key=period_sort_key):
        ws.append(
            [
                period["year"],
                period["month"],
                sheet_title(period, "Summary"),
                sheet_title(period, "Txn"),
                sheet_title(period, "Unparsed"),
            ]
        )

    years = sorted({period["year"] for period in periods_from_monthly_workbook(wb)}, key=lambda value: 9999 if value == "unknown" else int(value))
    ws.append(["", "", "", "", ""])
    ws.append(["Yearly Sheets", "", "", "", ""])
    ws.append(["Year", "Performance Sheet", "Category Sheet", "Transactions Sheet", ""])
    for year in years:
        ws.append([year, f"Yearly_{year}", f"Categories_{year}", f"AllTxn_{year}", ""])


def append_rows_to_monthly_workbook(wb, period, records, unparsed_records):
    txn_sheet = sheet_title(period, "Txn")
    unparsed_sheet = sheet_title(period, "Unparsed")
    summary_sheet = sheet_title(period, "Summary")

    ws_txn = ensure_report_sheet(wb, txn_sheet, TRANSACTION_HEADERS)
    ws_unparsed = ensure_report_sheet(wb, unparsed_sheet, UNPARSED_HEADERS)
    ws_summary = ensure_report_sheet(wb, summary_sheet)

    remove_transaction_totals(ws_txn)
    existing_uids = {ws_txn.cell(row=row, column=1).value for row in range(2, ws_txn.max_row + 1)}
    for record in records:
        if record["uid"] in existing_uids:
            continue
        p = record["parsed"]
        income = p["income"] or ""
        expense = p["expense"] or ""
        net = float(income or 0) - float(expense or 0)
        amount = p["amount"] or income or expense
        ws_txn.append(
            [
                record["uid"],
                p["transaction_date"],
                record.get("email_date", ""),
                p["type"],
                p["category"],
                p["party"],
                p["description"],
                amount,
                income,
                expense,
                net,
                p["source_account"],
                p["destination"],
                record["subject"],
            ]
        )

    existing_unparsed = {ws_unparsed.cell(row=row, column=1).value for row in range(2, ws_unparsed.max_row + 1)}
    for record in unparsed_records:
        if record["uid"] in existing_unparsed:
            continue
        ws_unparsed.append(
            [record["uid"], record["transaction_date"], record["party"], record["amount"], record["category"], record["subject"]]
        )

    add_transaction_totals(ws_txn)
    write_summary(ws_summary, ws_txn, period)


def group_by_period(records, unparsed_records):
    grouped = {}

    for record in records:
        period = transaction_period(record["parsed"]["transaction_date"])
        grouped.setdefault(period["key"], {"period": period, "records": [], "unparsed": []})
        grouped[period["key"]]["records"].append(record)

    for record in unparsed_records:
        period = transaction_period(record["transaction_date"])
        grouped.setdefault(period["key"], {"period": period, "records": [], "unparsed": []})
        grouped[period["key"]]["unparsed"].append(record)

    return grouped


def collect_transaction_sheet_totals(wb, period):
    ws = wb[sheet_title(period, "Txn")]
    income = 0
    expense = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "TOTAL":
            continue
        income += float(ws.cell(row=row, column=9).value or 0)
        expense += float(ws.cell(row=row, column=10).value or 0)
    return {"period": period, "income": income, "expense": expense, "net": income - expense}


def read_all_monthly_totals(wb):
    return [collect_transaction_sheet_totals(wb, period) for period in periods_from_monthly_workbook(wb)]


def category_totals_for_year(wb, year):
    totals = {}
    for item in read_all_monthly_totals(wb):
        if item["period"]["year"] != year:
            continue

        ws = wb[sheet_title(item["period"], "Txn")]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "TOTAL":
                continue
            category = reporting_category(ws.cell(row=row, column=5).value)
            income = float(ws.cell(row=row, column=9).value or 0)
            expense = float(ws.cell(row=row, column=10).value or 0)
            bucket = totals.setdefault(category, {"income": 0, "expense": 0})
            bucket["income"] += income
            bucket["expense"] += expense

    return totals


def remove_existing_yearly_sheets(wb):
    for sheet_name in list(wb.sheetnames):
        if sheet_name.startswith("Yearly_") or sheet_name.startswith("Categories_") or sheet_name.startswith("AllTxn_"):
            del wb[sheet_name]


def write_yearly_sheets(wb):
    remove_existing_yearly_sheets(wb)
    totals = read_all_monthly_totals(wb)
    years = sorted({item["period"]["year"] for item in totals}, key=lambda value: 9999 if value == "unknown" else int(value))

    for year in years:
        ws = wb.create_sheet(f"Yearly_{year}"[:31])

        same_year = sorted([item for item in totals if item["period"]["year"] == year], key=lambda item: period_sort_key(item["period"]))
        ws.append(["Year", year])
        ws.append(["Metric", "Amount"])
        ws.append(["Total Income", sum(item["income"] for item in same_year)])
        ws.append(["Total Expenses", sum(item["expense"] for item in same_year)])
        ws.append(["Net Income", sum(item["net"] for item in same_year)])
        ws.append(["", ""])
        ws.append(["Month", "Income", "Expenses", "Net"])
        for item in same_year:
            ws.append([item["period"]["month"], item["income"], item["expense"], item["net"]])

        ws_cat = wb.create_sheet(f"Categories_{year}"[:31])
        ws_cat.append(["Category", "Income", "Expenses", "Net"])
        for category, values in sorted(category_totals_for_year(wb, year).items()):
            ws_cat.append([category, values["income"], values["expense"], values["income"] - values["expense"]])

        ws_txn = wb.create_sheet(f"AllTxn_{year}"[:31])
        ws_txn.append(TRANSACTION_HEADERS)
        for item in same_year:
            source = wb[sheet_title(item["period"], "Txn")]
            for row in range(2, source.max_row + 1):
                if source.cell(row=row, column=1).value == "TOTAL":
                    continue
                ws_txn.append([source.cell(row=row, column=col).value for col in range(1, len(TRANSACTION_HEADERS) + 1)])
        add_transaction_totals(ws_txn)


def transaction_dict_from_record(record):
    p = record["parsed"]
    income = p["income"] or ""
    expense = p["expense"] or ""
    amount = p["amount"] or income or expense
    net = float(income or 0) - float(expense or 0)
    return refine_item_category({
        "uid": record["uid"],
        "transaction_date": p["transaction_date"],
        "email_date": record.get("email_date", ""),
        "type": p["type"],
        "category": p["category"],
        "party": p["party"],
        "description": p["description"],
        "amount": amount,
        "income": income,
        "expense": expense,
        "net": net,
        "source_account": p["source_account"],
        "destination": p["destination"],
        "subject": record["subject"],
    })


def transaction_dict_from_sheet_row(ws, row):
    return refine_item_category({
        "uid": ws.cell(row=row, column=1).value,
        "transaction_date": ws.cell(row=row, column=2).value,
        "email_date": ws.cell(row=row, column=3).value,
        "type": ws.cell(row=row, column=4).value,
        "category": ws.cell(row=row, column=5).value,
        "party": ws.cell(row=row, column=6).value,
        "description": ws.cell(row=row, column=7).value,
        "amount": ws.cell(row=row, column=8).value,
        "income": ws.cell(row=row, column=9).value,
        "expense": ws.cell(row=row, column=10).value,
        "net": ws.cell(row=row, column=11).value,
        "source_account": ws.cell(row=row, column=12).value,
        "destination": ws.cell(row=row, column=13).value,
        "subject": ws.cell(row=row, column=14).value,
    })


def transaction_dict_to_row(item):
    return [
        item.get("uid", ""),
        item.get("transaction_date", ""),
        item.get("email_date", ""),
        item.get("type", ""),
        item.get("category", ""),
        item.get("party", ""),
        item.get("description", ""),
        item.get("amount", ""),
        item.get("income", ""),
        item.get("expense", ""),
        item.get("net", ""),
        item.get("source_account", ""),
        item.get("destination", ""),
        item.get("subject", ""),
    ]


def collect_existing_transactions():
    seen_uids = set()
    items = []
    if not OUTPUT_DIR.exists():
        return items

    for path in sorted(OUTPUT_DIR.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            if not (ws.title.endswith("_Txn") or ws.title.startswith("AllTxn_")):
                continue
            for row in range(2, ws.max_row + 1):
                uid = ws.cell(row=row, column=1).value
                if uid in ("", None, "TOTAL") or uid in seen_uids:
                    continue
                seen_uids.add(uid)
                items.append(transaction_dict_from_sheet_row(ws, row))

    return items


def infer_missing_years(items, start_year=2026):
    def sort_key(item):
        uid = item.get("uid")
        uid_key = int(uid) if str(uid).isdigit() else 0
        return (email_sort_value(item.get("email_date")), uid_key)

    sorted_items = sorted(items, key=sort_key, reverse=True)
    current_year = start_year
    previous_month = None

    for item in sorted_items:
        explicit_year = transaction_explicit_year(item.get("transaction_date"))
        month_number = transaction_month_number(item.get("transaction_date"))

        if explicit_year:
            inferred_year = explicit_year
            current_year = explicit_year
        else:
            if previous_month and month_number and month_number > previous_month:
                current_year -= 1
            inferred_year = current_year

        item["report_year"] = str(inferred_year)
        item["report_month"] = report_month_for_item(item)

        if month_number:
            previous_month = month_number

    return sorted_items


def clean_generated_year_workbooks():
    for path in OUTPUT_DIR.glob("financial_report_*.xlsx"):
        if not path.name.startswith("~$") and "Dashboard" not in path.stem:
            path.unlink()


def dashboard_sheet_name(year):
    return f"Dashboard_{year}"[:31]


def top_dashboard_categories(monthly_category_totals):
    totals = {}
    for (_month, category), values in monthly_category_totals.items():
        totals[category] = totals.get(category, 0) + float(values.get("expense") or 0) + float(values.get("investment") or 0)
    categories = [category for category, _amount in sorted(totals.items(), key=lambda item: item[1], reverse=True) if category != "Income"]
    return (categories + ["Dining", "Investment", "Uncategorised Expense"])[:3]


def write_chart_helper(wb, year, selected_month, category_names):
    ws = wb.create_sheet("Chart_helper")
    ws.append([None, "Income", "Expenses", "Net income", "Saving rate", *category_names, None])
    ws.append(
        [
            f"='{dashboard_sheet_name(year)}'!J5",
            '=XLOOKUP($A2,Monthly_Summaries!$A$2:$A$13,Monthly_Summaries!$B$2:$B$13,"")',
            '=XLOOKUP($A2,Monthly_Summaries!$A$2:$A$13,Monthly_Summaries!$E$2:$E$13,"")',
            '=XLOOKUP($A2,Monthly_Summaries!$A$2:$A$13,Monthly_Summaries!$H$2:$H$13,"")',
            '=XLOOKUP($A2,Monthly_Summaries!$A$2:$A$13,Monthly_Summaries!$K$2:$K$13,"")',
            *[f'=SUMIFS(Categories!${"E" if category == "Investment" else "D"}:${"E" if category == "Investment" else "D"},Categories!$A:$A,$A2,Categories!$B:$B,{chr(70 + index)}$1)' for index, category in enumerate(category_names)],
            None,
        ]
    )
    ws.append([None, "Income change", "Expenses change", "Projected Emergency Funds(20%):", None, None, None, None, None])
    ws.append(
        [
            None,
            '=XLOOKUP($A2,Monthly_Summaries!$A$2:$A$13,Monthly_Summaries!$D$2:$D$13,"Nil")',
            '=XLOOKUP($A2,Monthly_Summaries!$A$2:$A$13,Monthly_Summaries!$F$2:$F$13,"Nil")',
            "=Yearly_Summary!D25",
            None,
            None,
            None,
            None,
            None,
        ]
    )


def write_dashboard(wb, year, ordered_months, category_names):
    ws = wb.create_sheet(dashboard_sheet_name(year), 0)
    ws["B2"] = "Jet's Financial Dashboard"
    ws["I5"] = "Select Month:"
    selected_month = ordered_months[-1][0] if ordered_months else "January"
    ws["J5"] = selected_month

    month_list = ",".join(month for month, _bucket in ordered_months) or "January"
    validation = DataValidation(type="list", formula1=f'"{month_list}"', allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(ws["J5"])

    performance_chart = BarChart()
    performance_chart.title = "Monthly Performance"
    performance_chart.y_axis.title = "SGD"
    performance_chart.x_axis.title = "Selected Month"
    performance_chart.add_data(Reference(wb["Chart_helper"], min_col=2, max_col=4, min_row=1, max_row=2), titles_from_data=True)
    performance_chart.set_categories(Reference(wb["Chart_helper"], min_col=1, min_row=2, max_row=2))
    performance_chart.height = 7
    performance_chart.width = 15
    ws.add_chart(performance_chart, "B7")

    category_chart = PieChart()
    category_chart.title = "Selected Month Spending"
    category_chart.add_data(Reference(wb["Chart_helper"], min_col=6, max_col=5 + len(category_names), min_row=2, max_row=2), from_rows=True)
    category_chart.set_categories(Reference(wb["Chart_helper"], min_col=6, max_col=5 + len(category_names), min_row=1, max_row=1))
    category_chart.dataLabels = DataLabelList()
    category_chart.dataLabels.showPercent = True
    category_chart.height = 7
    category_chart.width = 12
    ws.add_chart(category_chart, "L7")


def write_year_workbook(year, items):
    report_name = "financial_report_2025_Dashboard.xlsx" if str(year) == "2025" else f"financial_report_{year}.xlsx"
    path = OUTPUT_DIR / report_name
    wb = Workbook()
    wb.active.title = "Index"

    ws_index = wb["Index"]
    ws_index.append([f"Financial Report {year}", ""])
    ws_index.append(["Generated Sheets", ""])

    ws_year = wb.create_sheet("Yearly_Summary")
    total_income = sum(float(item.get("income") or 0) for item in items)
    total_expense = sum(float(item.get("expense") or 0) for item in items if not is_investment_category(item.get("category")))
    total_investment = sum(float(item.get("expense") or 0) for item in items if is_investment_category(item.get("category")))
    ws_year.append(["Year", year])
    ws_year.append(["Metric", "Amount"])
    ws_year.append(["Total Income", total_income])
    ws_year.append(["Total Expenses", total_expense])
    ws_year.append(["Total Investments", total_investment])
    ws_year.append(["Net Income", total_income - total_expense])
    ws_year.append(["Net After Investments", total_income - total_expense - total_investment])
    ws_year.append(["", ""])
    ws_year.append(
        [
            "Month",
            "Income",
            "Income % Change",
            "Expenses",
            "Expense % Change",
            "Investments",
            "Net Income",
            "Net Income % Change",
            "Net After Investments",
        ]
    )

    by_month = {}
    for item in items:
        month = item.get("report_month") or "unknown"
        bucket = by_month.setdefault(month, {"income": 0, "expense": 0, "investment": 0, "items": []})
        bucket["income"] += float(item.get("income") or 0)
        if is_investment_category(item.get("category")):
            bucket["investment"] += float(item.get("expense") or 0)
        else:
            bucket["expense"] += float(item.get("expense") or 0)
        bucket["items"].append(item)

    ordered_months = sorted(by_month.items(), key=lambda pair: MONTH_ORDER.get(pair[0], 99))
    previous_bucket = None
    for month, bucket in ordered_months:
        net_income = bucket["income"] - bucket["expense"]
        previous_net_income = previous_bucket["income"] - previous_bucket["expense"] if previous_bucket else ""
        ws_year.append(
            [
                month,
                bucket["income"],
                percent_change(bucket["income"], previous_bucket["income"]) if previous_bucket else "",
                bucket["expense"],
                percent_change(bucket["expense"], previous_bucket["expense"]) if previous_bucket else "",
                bucket["investment"],
                net_income,
                percent_change(net_income, previous_net_income) if previous_bucket else "",
                net_income - bucket["investment"],
            ]
        )
        previous_bucket = bucket

    ws_monthly = wb.create_sheet("Monthly_Summaries")
    ws_monthly.append(
        [
            "Month",
            "True Income",
            "Fake income",
            "Income % Change",
            "Expenses",
            "Expense % Change",
            "Investments",
            "True Net Income",
            "Fake Net Income",
            "True Saving rate",
            "Fake Saving rate",
            "Net Income % Change",
            "Net After Investments",
            "Transaction Count",
        ]
    )
    for row_number, (month, bucket) in enumerate(ordered_months, start=2):
        ws_monthly.append(
            [
                month,
                bucket["income"],
                f"=B{row_number}+2000",
                excel_percent_change_formula(f"A{row_number}", f"B{row_number}", f"B{row_number - 1}"),
                bucket["expense"],
                excel_percent_change_formula(f"A{row_number}", f"E{row_number}", f"E{row_number - 1}"),
                bucket["investment"],
                f"=B{row_number}-E{row_number}",
                f"=C{row_number}-E{row_number}",
                f'=IF(B{row_number}=0,"Nil",H{row_number}/B{row_number})',
                f'=IF(C{row_number}=0,"Nil",I{row_number}/C{row_number})',
                excel_percent_change_formula(f"A{row_number}", f"H{row_number}", f"H{row_number - 1}"),
                f"=H{row_number}-G{row_number}",
                len(bucket["items"]),
            ]
        )

    ws_cat = wb.create_sheet("Categories")
    ws_cat.append(["Month", "Category", "Income", "Expenses", "Investments", "Net", "% of Month Expenses", "% of Year Expenses"])
    category_totals = {}
    monthly_category_totals = {}
    for item in items:
        category = reporting_category(item.get("category"))
        month = item.get("report_month") or "unknown"
        bucket = category_totals.setdefault(category, {"income": 0, "expense": 0, "investment": 0})
        monthly_bucket = monthly_category_totals.setdefault((month, category), {"income": 0, "expense": 0, "investment": 0})
        bucket["income"] += float(item.get("income") or 0)
        monthly_bucket["income"] += float(item.get("income") or 0)
        if is_investment_category(item.get("category")):
            bucket["investment"] += float(item.get("expense") or 0)
            monthly_bucket["investment"] += float(item.get("expense") or 0)
        else:
            bucket["expense"] += float(item.get("expense") or 0)
            monthly_bucket["expense"] += float(item.get("expense") or 0)

    year_expense_total = sum(values["expense"] for values in category_totals.values())
    for month, bucket in ordered_months:
        month_expense_total = bucket["expense"]
        for (category_month, category), values in sorted(monthly_category_totals.items(), key=lambda pair: (MONTH_ORDER.get(pair[0][0], 99), pair[0][1])):
            if category_month != month:
                continue
            expense_share_month = values["expense"] / month_expense_total if month_expense_total else ""
            expense_share_year = values["expense"] / year_expense_total if year_expense_total else ""
            ws_cat.append(
                [
                    month,
                    category,
                    values["income"],
                    values["expense"],
                    values["investment"],
                    values["income"] - values["expense"] - values["investment"],
                    expense_share_month,
                    expense_share_year,
                ]
            )

    ws_all = wb.create_sheet("All_Transactions")
    ws_all.append(TRANSACTION_HEADERS)
    for item in sorted(items, key=lambda row: int(row["uid"]) if str(row.get("uid")).isdigit() else 0, reverse=True):
        ws_all.append(transaction_dict_to_row(item))
    add_transaction_totals(ws_all)

    ws_index.append(["", ""])
    ws_index.append(["Yearly_Summary", "Overall yearly performance"])
    ws_index.append(["Monthly_Summaries", "All monthly summary rows in one sheet"])
    ws_index.append(["Categories", "Monthly income, expenses, investments, and spending share by category"])
    ws_index.append(["All_Transactions", "All income and expenses for this year"])

    for month, bucket in ordered_months:
        period = period_for_year_month(year, month)
        ws_txn = wb.create_sheet(sheet_title(period, "Txn"))
        ws_txn.append(TRANSACTION_HEADERS)
        for item in sorted(bucket["items"], key=lambda row: int(row["uid"]) if str(row.get("uid")).isdigit() else 0, reverse=True):
            ws_txn.append(transaction_dict_to_row(item))
        add_transaction_totals(ws_txn)

        ws_summary = wb.create_sheet(sheet_title(period, "Summary"))
        write_summary(ws_summary, ws_txn, period)
        ws_index.append([ws_txn.title, f"{month} transactions"])
        ws_index.append([ws_summary.title, f"{month} summary"])

    apply_template(wb)
    wb.save(path)
    return str(path)


def rebuild_year_workbooks(new_records=None):
    OUTPUT_DIR.mkdir(exist_ok=True)
    by_uid = {}
    for item in collect_existing_transactions():
        by_uid[item["uid"]] = item

    for record in new_records or []:
        by_uid[record["uid"]] = transaction_dict_from_record(record)

    inferred_items = infer_missing_years(list(by_uid.values()), start_year=2026)
    by_year = {}
    for item in inferred_items:
        by_year.setdefault(item["report_year"], []).append(item)

    clean_generated_year_workbooks()
    return [write_year_workbook(year, items) for year, items in sorted(by_year.items(), reverse=True)]


def append_rows(records, unparsed_records):
    return rebuild_year_workbooks(records)


def run_transaction_export():
    records, unparsed_records, max_uid_seen = fetch_new_bank_emails()
    if records or unparsed_records:
        updated_files = append_rows(records, unparsed_records)
        save_last_uid(max_uid_seen)
    else:
        updated_files = rebuild_year_workbooks()

    total_income = sum(float(record["parsed"]["income"] or 0) for record in records)
    total_expense = sum(float(record["parsed"]["expense"] or 0) for record in records)
    return {
        "transactions": len(records),
        "unparsed": len(unparsed_records),
        "income": total_income,
        "expense": total_expense,
        "net": total_income - total_expense,
        "files": updated_files,
    }


if __name__ == "__main__":
    summary = run_transaction_export()
    print(
        "Processed "
        f"{summary['transactions']} transaction(s), {summary['unparsed']} unparsed email(s). "
        f"Income SGD {summary['income']:.2f}, expenses SGD {summary['expense']:.2f}, net SGD {summary['net']:.2f}."
    )
    if summary["files"]:
        print("Updated workbook(s):")
        for file_path in summary["files"]:
            print(f"- {file_path}")
    else:
        print("No workbook updates.")
